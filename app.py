"""Flask entry point for the Keeper League dashboard."""

import os
import json
import logging
import subprocess

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory
from flask_login import current_user, login_required
from flask_wtf.csrf import CSRFProtect
from werkzeug.utils import secure_filename

from flask_socketio import SocketIO

import config
from models.database import db, init_db

csrf = CSRFProtect()
from models.auth import login_manager
from models.player import load_players_csv
from models.draft_model import rank_players, factor_breakdown, compute_historical_draft_scores
from scrapers.footywire import (
    build_master_player_list, load_player_sc_history,
)
from scrapers.csv_import import import_players_csv, import_sc_scores_csv
from scrapers.stats_loader import (
    load_player_detailed_stats,
    backfill_sc_from_fitzroy,
    load_player_sc_history_fitzroy,
)


_XLSX_PATH = os.environ.get("RATINGS_XLSX_PATH",
    os.path.join(os.path.expanduser("~"), "OneDrive", "Documents", "AFL 2025.1.xlsx")
)

# Team name mapping: Excel name -> DB name
_XLSX_TEAM_MAP = {
    "Brisbane": "Brisbane Lions",
    "Greater Western Sydney": "GWS",
    "St. Kilda": "St Kilda",
}

# Exact name overrides: DB name -> XLSX name (when names are just different)
_NAME_OVERRIDES = {
    ("Michael Frederick", "Fremantle"): "Minairo Frederick",
    ("Max King", "Sydney"): "Maxy King",
    ("Nikolas Cox", "Essendon"): "Nik Cox",
    ("Jacob Farrow", "Essendon"): "Jackson Farrow",
    ("Indy Cotton", "Adelaide"): "Indi Cotton",
    ("Hussien El Achkar", "Essendon"): "Hussein El Achkar",
    ("Jack Hutchinson", "West Coast"): "Jack Hutchison",
    ("Maurice Rioli", "Richmond"): "Maurice Rioli Jr",
    ("Robert Hansen", "North Melbourne"): "Robert Hansen Jr",
}

# Bidirectional nickname map for fuzzy player matching
_NICKNAME_MAP = {
    "brad": "bradley", "cam": "cameron", "dan": "daniel",
    "ed": "edward", "harry": "harrison", "josh": "joshua",
    "jake": "jacob", "tom": "thomas", "will": "william",
    "matt": "matthew", "ben": "benjamin", "sam": "samuel",
    "nick": "nicholas", "mitch": "mitchell", "alex": "alexander",
    "tim": "timothy", "rob": "robert", "pat": "patrick",
    "joe": "joseph", "charlie": "charles", "liam": "william",
    "zac": "zachary", "zach": "zachary", "nat": "nathan",
    "nic": "nicholas", "mike": "michael", "chris": "christopher",
    "jim": "james", "jon": "jonathan", "dave": "david",
    "lachlan": "lachie", "ollie": "oliver", "wil": "will",
    "willem": "will", "jack": "jackson", "jaime": "jamie",
}
# Build reverse map (long -> short)
_NICKNAME_MAP_REV = {}
for short, long in _NICKNAME_MAP.items():
    _NICKNAME_MAP_REV.setdefault(long, []).append(short)


def _name_variants(name: str):
    """Generate plausible first-name variants for a full name."""
    parts = name.strip().split()
    if len(parts) < 2:
        return []
    first = parts[0].lower()
    rest = " ".join(parts[1:])
    variants = []
    # short -> long
    if first in _NICKNAME_MAP:
        variants.append(f"{_NICKNAME_MAP[first].title()} {rest}")
    # long -> short(s)
    if first in _NICKNAME_MAP_REV:
        for short in _NICKNAME_MAP_REV[first]:
            variants.append(f"{short.title()} {rest}")
    return variants


def _sync_ratings_to_db(app):
    """Import rating & potential from AFL 2025.1.xlsx into the afl_player table."""
    from models.database import AflPlayer, RatingLog

    if not os.path.exists(_XLSX_PATH):
        app.logger.info("Ratings sync: XLSX not found, skipping")
        return

    # Single-worker guard. Gunicorn spawns N workers and each calls create_app(),
    # so without a lock every worker would re-run the sync and write duplicate
    # rating_log rows. Non-blocking flock — first worker wins, others skip.
    # Kernel releases on process exit, so no stale-lock recovery needed.
    try:
        import fcntl
    except ImportError:
        fcntl = None  # Windows dev — single-process, no guard needed

    lock_file = None
    if fcntl is not None:
        try:
            lock_file = open(_XLSX_PATH + ".sync.lock", "w")
            fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        except (BlockingIOError, OSError):
            if lock_file is not None:
                lock_file.close()
            app.logger.info("Ratings sync: another worker is syncing, skipping")
            return

    try:
        with app.app_context():
            # Staleness check
            import datetime as _dt
            xlsx_mtime = _dt.datetime.fromtimestamp(os.path.getmtime(_XLSX_PATH))
            newest = AflPlayer.query.filter(AflPlayer.rating.isnot(None)).order_by(
                AflPlayer.updated_at.desc()
            ).first()
            if newest and newest.updated_at:
                db_time = newest.updated_at.replace(tzinfo=None)
                if db_time >= xlsx_mtime:
                    app.logger.info("Ratings sync: XLSX not newer than DB, skipping")
                    return

            import openpyxl
            wb = openpyxl.load_workbook(_XLSX_PATH, read_only=True, data_only=True)
            ws = wb["Player Database"]

            # Find the current-year column for start-of-year rating
            # After Dec 1 use next year's column; before Dec 1 use current year
            from datetime import date
            today = date.today()
            target_year = today.year + 1 if today.month == 12 else today.year
            headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            year_col_idx = None
            for ci, h in enumerate(headers):
                if h and str(h).strip().isdigit() and int(str(h)) == target_year:
                    year_col_idx = ci
                    break

            # Read rows: col 0=Player, 2=Rating, 3=Potential, 4=Team, year_col_idx=start-of-year
            xlsx_players = []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                name = row[0]
                if not name or not isinstance(name, str):
                    continue
                team_raw = row[4] or ""
                team = _XLSX_TEAM_MAP.get(team_raw, team_raw)
                rating = row[2] if isinstance(row[2], (int, float)) else None
                potential = row[3] if isinstance(row[3], (int, float)) else None
                rating_start = None
                if year_col_idx is not None and len(row) > year_col_idx:
                    val = row[year_col_idx]
                    if isinstance(val, (int, float)):
                        rating_start = int(val)
                if rating is not None:
                    rating = int(rating)
                    if rating < 0 or rating > 100:
                        logger.warning("Rating %d out of range for %s — skipping", rating, name)
                        continue
                if potential is not None:
                    potential = int(potential)
                    if potential < 0 or potential > 100:
                        logger.warning("Potential %d out of range for %s — skipping", potential, name)
                        continue
                xlsx_players.append((name.strip(), team.strip(), rating, potential, rating_start))
            wb.close()

            # Build DB lookup: (lowercase name, team) -> AflPlayer
            all_db = AflPlayer.query.all()
            db_lookup = {}
            for ap in all_db:
                key = (ap.name.lower(), ap.afl_team)
                db_lookup[key] = ap

            # Build reverse override map: (xlsx_name_lower, team) -> db AflPlayer
            override_lookup = {}
            for (db_name, db_team), xlsx_name in _NAME_OVERRIDES.items():
                ap = db_lookup.get((db_name.lower(), db_team))
                if ap:
                    override_lookup[(xlsx_name.lower(), db_team)] = ap

            matched = 0
            unmatched = []

            for xlsx_name, xlsx_team, rating, potential, rating_start in xlsx_players:
                # Pass 1: exact match on (name, team)
                key = (xlsx_name.lower(), xlsx_team)
                ap = db_lookup.get(key)

                # Pass 2: explicit overrides
                if ap is None:
                    ap = override_lookup.get(key)

                # Pass 3: nickname variants (both directions)
                if ap is None:
                    for variant in _name_variants(xlsx_name):
                        key2 = (variant.lower(), xlsx_team)
                        ap = db_lookup.get(key2)
                        if ap:
                            break

                if ap:
                    # Log changes before applying
                    rating_changed = (ap.rating != rating) if rating is not None else False
                    potential_changed = (ap.potential != potential) if potential is not None else False
                    if rating_changed or potential_changed:
                        log = RatingLog(
                            player_id=ap.id,
                            old_rating=ap.rating,
                            new_rating=rating,
                            old_potential=ap.potential,
                            new_potential=potential,
                            rating_start=rating_start if rating_start is not None else ap.rating_start,
                        )
                        db.session.add(log)
                    ap.rating = rating
                    ap.potential = potential
                    if rating_start is not None:
                        ap.rating_start = rating_start
                    matched += 1
                else:
                    unmatched.append((xlsx_name, xlsx_team))

            db.session.commit()
            app.logger.info(
                f"Ratings sync: {matched} players updated, "
                f"{len(unmatched)} unmatched from XLSX"
            )
            if unmatched and len(unmatched) <= 30:
                for n, t in unmatched:
                    app.logger.debug(f"  Unmatched: {n} ({t})")
    finally:
        if lock_file is not None:
            try:
                fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)
            except OSError:
                pass
            lock_file.close()


def _sync_players_to_db(app, force=False):
    """Bulk-import players.csv into the afl_player table if it's empty or stale.

    Args:
        force: If True, bypass staleness check and always sync.
    """
    from models.database import AflPlayer
    from models.player import load_players_csv, player_to_orm
    from models.draft_model import rank_players

    with app.app_context():
        csv_path = os.path.join(config.DATA_DIR, "players.csv")
        if not os.path.exists(csv_path):
            return

        db_count = AflPlayer.query.count()
        if not force and db_count > 0:
            # Check staleness: compare CSV mod-time to newest updated_at
            import datetime
            csv_mtime = datetime.datetime.fromtimestamp(os.path.getmtime(csv_path))
            newest = (
                AflPlayer.query
                .order_by(AflPlayer.updated_at.desc())
                .first()
            )
            if newest and newest.updated_at:
                # Strip timezone info for comparison (SQLite stores naive datetimes)
                db_time = newest.updated_at.replace(tzinfo=None)
                if db_time >= csv_mtime:
                    return  # DB is up-to-date

        players = load_players_csv()
        if not players:
            return

        # Rank them so draft_score is populated
        rank_players(players, config.DRAFT_WEIGHTS)

        # Upsert: match on (name, afl_team)
        existing = {(p.name, p.afl_team): p for p in AflPlayer.query.all()}
        added = 0
        updated = 0
        for p in players:
            orm_data = player_to_orm(p)
            key = (orm_data["name"], orm_data["afl_team"])
            if key in existing:
                row = existing[key]
                # Skip sc_avg/sc_avg_prev/games_played — these are now
                # computed from live PlayerStat data, not from CSV.
                _LIVE_COLS = {"name", "afl_team", "sc_avg", "sc_avg_prev", "games_played"}
                for col, val in orm_data.items():
                    if col not in _LIVE_COLS:
                        setattr(row, col, val)
                updated += 1
            else:
                row = AflPlayer(**orm_data)
                db.session.add(row)
                added += 1

        db.session.commit()
        total = AflPlayer.query.count()
        app.logger.info(f"Player sync: {added} added, {updated} updated, {total} total in DB")


def create_app():
    app = Flask(__name__)
    app.secret_key = config.SECRET_KEY
    app.config["SQLALCHEMY_DATABASE_URI"] = config.SQLALCHEMY_DATABASE_URI
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = config.SQLALCHEMY_TRACK_MODIFICATIONS
    app.config["SESSION_COOKIE_HTTPONLY"] = config.SESSION_COOKIE_HTTPONLY
    app.config["SESSION_COOKIE_SAMESITE"] = config.SESSION_COOKIE_SAMESITE
    app.config["SESSION_COOKIE_SECURE"] = config.SESSION_COOKIE_SECURE
    app.config["MAX_CONTENT_LENGTH"] = config.MAX_CONTENT_LENGTH

    # Email config
    app.config["MAIL_SERVER"] = config.MAIL_SERVER
    app.config["MAIL_PORT"] = config.MAIL_PORT
    app.config["MAIL_USE_TLS"] = config.MAIL_USE_TLS
    app.config["MAIL_USERNAME"] = config.MAIL_USERNAME
    app.config["MAIL_PASSWORD"] = config.MAIL_PASSWORD
    app.config["MAIL_DEFAULT_SENDER"] = config.MAIL_DEFAULT_SENDER

    # Init extensions
    init_db(app)
    login_manager.init_app(app)
    csrf.init_app(app)

    # Flask-Mail (lazy — only sends if MAIL_USERNAME is configured)
    from flask_mail import Mail
    mail = Mail(app)
    app.extensions["mail"] = mail

    # Enable SQLite WAL mode for concurrent reads
    if "sqlite" in config.SQLALCHEMY_DATABASE_URI:
        from sqlalchemy import event as sa_event, engine as sa_engine

        @sa_event.listens_for(sa_engine.Engine, "connect")
        def _set_sqlite_pragma(dbapi_conn, connection_record):
            cursor = dbapi_conn.cursor()
            cursor.execute("PRAGMA journal_mode=WAL")
            cursor.execute("PRAGMA busy_timeout=30000")
            cursor.close()

    # Sync CSV players into SQLite, then ratings from XLSX
    _sync_players_to_db(app)
    _sync_ratings_to_db(app)

    # Register blueprints
    from blueprints.auth import auth_bp
    from blueprints.leagues import leagues_bp
    from blueprints.draft import draft_bp
    from blueprints.team import team_bp
    from blueprints.trades import trades_bp
    from blueprints.matchups import matchups_bp
    from blueprints.admin import admin_bp
    from blueprints.comms import comms_bp
    from blueprints.reserve7s import reserve7s_bp
    app.register_blueprint(auth_bp)
    app.register_blueprint(leagues_bp)
    app.register_blueprint(draft_bp)
    csrf.exempt(draft_bp)   # draft uses SocketIO (own origin check) + @login_required
    csrf.exempt(team_bp)    # 15+ AJAX API endpoints; all behind @login_required + team ownership
    csrf.exempt(leagues_bp) # SPA sends JSON, not form tokens; all behind @login_required
    app.register_blueprint(team_bp)
    app.register_blueprint(trades_bp)
    csrf.exempt(trades_bp)  # SPA sends JSON; all behind @login_required
    app.register_blueprint(matchups_bp)
    csrf.exempt(matchups_bp)  # SPA sends JSON; sync-scores etc behind @login_required
    app.register_blueprint(admin_bp)
    csrf.exempt(admin_bp)   # SPA sends JSON; all behind @login_required + is_admin
    app.register_blueprint(comms_bp)
    csrf.exempt(comms_bp)   # SPA sends JSON; all behind @login_required
    app.register_blueprint(reserve7s_bp)
    csrf.exempt(reserve7s_bp)  # 7s team API uses @login_required + ownership checks

    # SPA API blueprint (JSON endpoints for React frontend)
    from blueprints.spa_api import spa_api
    app.register_blueprint(spa_api)
    csrf.exempt(spa_api)  # all JSON API; behind @login_required
    csrf.exempt(auth_bp)  # auth API endpoints need CSRF exemption for JSON login

    # SPA directory for React frontend
    _spa_dir = os.path.join(app.static_folder, "spa")
    _spa_mode = os.environ.get("SPA_MODE", "0") == "1"

    def _send_spa_shell():
        """Serve the SPA index.html with strict no-cache headers.
        The HTML embeds cache-busted asset URLs (style.css?v=<hash>);
        if THIS file caches, the busted URLs go stale and users see
        old CSS even after a deploy. We discovered the user's browser
        was caching the shell for 30 min behind nginx — toast redesign
        appeared not to render because the cached HTML still referenced
        the pre-bump v=20260413e string. No-store closes that loop."""
        resp = send_from_directory(_spa_dir, "index.html")
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp

    # ── SPA-mode: intercept browser navigations and serve the React shell ──
    # Only catches real page navigations (user clicking link / typing URL).
    # Fetch/XHR API calls pass through to Flask as normal.
    if _spa_mode:
        @app.before_request
        def _spa_intercept():
            if request.method != "GET":
                return None

            path = request.path

            # Never intercept static files
            if path.startswith("/static/"):
                return None

            # Never intercept any API-like path
            if path.startswith("/api/") or "/api/" in path:
                return None
            if request.args.get("format") == "json":
                return None
            if path.startswith("/auth/api/"):
                return None
            if path.startswith("/push/"):
                return None
            # Player profile page is still Jinja2 — let it through
            if path.startswith("/player/"):
                return None
            # Admin hub pages are still Jinja2 — let them through
            if path.startswith("/admin/") or path == "/admin":
                return None

            # ── Key check: only intercept BROWSER NAVIGATIONS ──
            # Browsers send Sec-Fetch-Mode: navigate for page loads.
            # fetch() sends cors/same-origin/no-cors. This is the
            # reliable way to distinguish page nav from JS API calls
            # (Accept: */* matches html so we can't rely on that).
            sfm = request.headers.get("Sec-Fetch-Mode", "")
            if sfm and sfm != "navigate":
                return None

            # Fallback for older browsers that don't send Sec-Fetch-Mode:
            # Check if Accept explicitly prefers html over json.
            if not sfm:
                accept = request.headers.get("Accept", "")
                # fetch() default is */* — don't intercept that
                if "*/*" in accept and "text/html" not in accept:
                    return None
                if "application/json" in accept:
                    return None
                if "text/html" not in accept:
                    return None

            return _send_spa_shell()

        # Return 401 JSON (not 302 redirect) when @login_required fails.
        # In SPA mode, ALL requests are either page navigations (already
        # intercepted above) or API/fetch calls. So any request reaching
        # this handler is an API call and should get 401 JSON.
        @login_manager.unauthorized_handler
        def _api_unauthorized():
            return jsonify({"error": "Authentication required"}), 401

    # Legacy /spa/ path — redirect to root so old bookmarks work
    @app.route("/spa/", defaults={"path": ""})
    @app.route("/spa/<path:path>")
    def spa_catchall(path=""):
        """Redirect legacy /spa/* URLs to their root equivalents."""
        if _spa_mode:
            target = f"/{path}" if path else "/"
            qs = request.query_string.decode()
            if qs:
                target += f"?{qs}"
            return redirect(target, 301)
        return _send_spa_shell()

    # Static asset cache buster (hash of style.css mtime)
    import hashlib
    _css_path = os.path.join(app.static_folder, "style.css")
    try:
        _asset_v = hashlib.md5(str(os.path.getmtime(_css_path)).encode()).hexdigest()[:8]
    except Exception:
        _asset_v = "1"

    # Tiny Jinja filter — convert "#58a6ff" → "88,166,255" so templates
    # can drive --css-var: rgb(R,G,B) and rgba(R,G,B, alpha) gradients
    # without a separate context var per colour.
    @app.template_filter("hex_to_rgb")
    def _hex_to_rgb(hex_str):
        if not hex_str or not isinstance(hex_str, str):
            return "88,166,255"
        h = hex_str.strip().lstrip("#")
        if len(h) == 3:
            h = "".join(c * 2 for c in h)
        if len(h) != 6:
            return "88,166,255"
        try:
            return f"{int(h[0:2], 16)},{int(h[2:4], 16)},{int(h[4:6], 16)}"
        except ValueError:
            return "88,166,255"

    # Context processor — inject globals into all templates
    @app.context_processor
    def inject_globals():
        ctx = {
            "TEAM_LOGOS": config.TEAM_LOGOS,
            "TEAM_COLOURS": config.TEAM_COLOURS,
            "TEAM_ABBR": config.TEAM_ABBR,
            "SCORING_TYPE_LABELS": config.SCORING_TYPE_LABELS,
            "ASSET_V": _asset_v,
        }
        if current_user.is_authenticated:
            from models.database import FantasyTeam
            from models.notification_manager import get_unread_count
            user_teams = FantasyTeam.query.filter_by(
                owner_id=current_user.id
            ).all()
            ctx["user_leagues"] = [(t.league, t) for t in user_teams]
            ctx["unread_notif_count"] = get_unread_count(current_user.id)
            ctx["theme_pref"] = getattr(current_user, "theme_preference", "dark") or "dark"
            ctx["has_completed_onboarding"] = getattr(current_user, "has_completed_onboarding", True)

            # If in a league-scoped route, provide nav helpers
            league_id = (request.view_args or {}).get("league_id")
            if league_id:
                nav_team = next(
                    (t for t in user_teams if t.league_id == league_id), None
                )
                ctx["nav_user_team"] = nav_team
                # Finals config for subnav visibility
                from models.database import League as _League, SeasonConfig as _SC, DraftSession
                _lg = db.session.get(_League, league_id)
                if _lg:
                    _sc = _SC.query.filter_by(league_id=league_id, year=_lg.season_year).first()
                    ctx["nav_finals_teams"] = _sc.finals_teams if _sc else 4
                    # Active draft session for nav link visibility
                    ctx["nav_active_draft"] = DraftSession.query.filter_by(
                        league_id=league_id, is_mock=False
                    ).filter(
                        DraftSession.status.in_(["scheduled", "in_progress", "paused"])
                    ).first() is not None
                    # Commissioner nav helpers
                    if _lg.commissioner_id == current_user.id:
                        ctx["nav_is_commissioner"] = True
                        from models.database import LongTermInjury as _LTI
                        ctx["pending_ltil_count"] = _LTI.query.filter_by(
                            league_id=league_id, removed_at=None, status="pending"
                        ).count()
        return ctx

    # ── Security headers ────────────────────────────────────────────
    # ── Global POST rate limiter ──────────────────────────────────
    from collections import defaultdict
    import time as _time
    _post_limiter = defaultdict(list)  # IP -> [timestamps]
    _POST_LIMIT = 30  # max POST requests per window
    _POST_WINDOW = 60  # seconds

    @app.before_request
    def rate_limit_posts():
        if request.method != "POST":
            return
        ip = request.remote_addr or "unknown"
        now = _time.time()
        _post_limiter[ip] = [t for t in _post_limiter[ip] if now - t < _POST_WINDOW]
        if len(_post_limiter[ip]) >= _POST_LIMIT:
            return jsonify({"error": "Rate limit exceeded"}), 429
        _post_limiter[ip].append(now)

    @app.after_request
    def add_security_headers(response):
        response.headers["X-Frame-Options"] = "SAMEORIGIN"
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
        response.headers["Content-Security-Policy"] = (
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline' cdn.jsdelivr.net cdnjs.cloudflare.com; "
            "style-src 'self' 'unsafe-inline' cdn.jsdelivr.net fonts.googleapis.com; "
            "font-src 'self' fonts.gstatic.com cdn.jsdelivr.net; "
            "img-src 'self' data: https: blob:; "
            "connect-src 'self' wss: ws:; "
            "frame-ancestors 'self'; "
            "worker-src 'self' blob:;"
        )
        # Stop browsers caching dynamic HTML so deploys show up immediately.
        # nginx serves /static/* directly with its own Cache-Control headers,
        # so this only affects Flask-rendered responses.
        if not request.path.startswith("/static/"):
            ctype = response.headers.get("Content-Type", "")
            if ctype.startswith("text/html") or ctype.startswith("application/json"):
                response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
                response.headers["Pragma"] = "no-cache"
                response.headers["Expires"] = "0"
        return response

    # ── Error handlers ────────────────────────────────────────────────
    @app.errorhandler(404)
    def page_not_found(e):
        if _spa_mode and request.accept_mimetypes.accept_html:
            return _send_spa_shell(), 200
        return render_template("errors/404.html"), 404

    @app.errorhandler(500)
    def internal_error(e):
        db.session.rollback()
        return render_template("errors/500.html"), 500

    @app.errorhandler(403)
    def forbidden(e):
        if _spa_mode and request.accept_mimetypes.accept_html:
            return _send_spa_shell(), 200
        return render_template("errors/403.html"), 403

    # ── Request logging (analytics) ──────────────────────────────────
    # Only count genuine page navigations, not polling/API/bot probes.
    _BOT_PATH_PREFIXES = (
        "/wp-admin", "/wp-login", "/wordpress", "/.env", "/.git",
        "/phpmyadmin", "/admin.php", "/vendor/", "/cgi-bin/",
    )
    _ANALYTICS_SKIP_PREFIXES = (
        "/static/", "/socket.io/", "/sw.js", "/manifest.json",
        "/push/", "/auth/api/",
    )

    @app.after_request
    def log_page_view(response):
        """Record genuine HTML page navigations to PageView for analytics."""
        path = request.path
        # Method filter — only GETs count
        if request.method != "GET":
            return response
        # Skip asset-like paths
        if path == "/favicon.ico" or path.endswith((".js", ".css", ".png", ".svg", ".ico", ".json", ".map", ".woff", ".woff2")):
            return response
        # Skip static / ws / api / push
        if any(path.startswith(p) for p in _ANALYTICS_SKIP_PREFIXES):
            return response
        # Skip any path with /api/ anywhere (catches /leagues/X/draft/api/foo)
        if "/api/" in path:
            return response
        # Skip bot-probe paths
        if any(path.startswith(p) for p in _BOT_PATH_PREFIXES):
            return response
        # Skip non-success responses (4xx/5xx) — login bounces, errors, bot 404s
        if not (200 <= response.status_code < 400):
            return response
        # Only log responses that are actually HTML pages
        ctype = response.headers.get("Content-Type", "")
        if not ctype.startswith("text/html"):
            return response
        try:
            import hashlib
            from models.database import PageView
            ip_raw = request.remote_addr or "unknown"
            ip_hash = hashlib.sha256(ip_raw.encode()).hexdigest()
            pv = PageView(
                user_id=current_user.id if current_user.is_authenticated else None,
                path=path,
                method=request.method,
                status_code=response.status_code,
                user_agent=(request.user_agent.string or "")[:256],
                ip_hash=ip_hash,
            )
            db.session.add(pv)
            db.session.commit()
        except Exception:
            db.session.rollback()
        return response

    # ── Helpers ──────────────────────────────────────────────────────

    # Process-local cache for ranked players. _get_ranked_players is
    # called on every /player/<name> hit and rank_players costs ~0.3s
    # for ~800 players. Output is keyed on (player_count, max_updated_at)
    # so any DB write to AflPlayer invalidates automatically without
    # needing an explicit invalidation hook.
    _ranked_cache = {"key": None, "players": None}

    def _get_ranked_players():
        from models.database import AflPlayer, db as _db
        from models.player import orm_to_player
        from sqlalchemy import func as _sa_func
        # Cheap key: row count + max updated_at. Both cost <1ms in SQLite.
        try:
            row = _db.session.query(
                _sa_func.count(AflPlayer.id),
                _sa_func.max(AflPlayer.updated_at) if hasattr(AflPlayer, "updated_at") else _sa_func.max(AflPlayer.id),
            ).first()
            cache_key = (row[0], row[1])
        except Exception:
            cache_key = None
        if cache_key and _ranked_cache["key"] == cache_key and _ranked_cache["players"] is not None:
            return _ranked_cache["players"]

        afl_players = AflPlayer.query.all()
        if afl_players:
            players = [orm_to_player(ap) for ap in afl_players]
            rank_players(players, config.DRAFT_WEIGHTS)
        else:
            # Fallback to CSV if DB is empty
            players = load_players_csv()
            if players:
                rank_players(players, config.DRAFT_WEIGHTS)
        if cache_key:
            _ranked_cache["key"] = cache_key
            _ranked_cache["players"] = players
        return players

    # ── Public routes (no login required) ────────────────────────────

    @app.route("/")
    def index():
        if not current_user.is_authenticated:
            return redirect(url_for("auth.login"))
        return redirect(url_for("leagues.league_list"))

    @app.route("/player/<name>")
    def player_detail(name):
        players = _get_ranked_players()
        player = next((p for p in players if p.name == name), None)
        if player is None:
            flash(f"Player '{name}' not found.", "warning")
            return redirect(url_for("leagues.league_list"))

        breakdown = factor_breakdown(player, players, config.DRAFT_WEIGHTS)

        # Load multi-year SC history (fitzRoy preferred — has all years; footywire fallback)
        sc_history = load_player_sc_history_fitzroy(name)
        if not sc_history.get("yearly_averages"):
            sc_history = load_player_sc_history(name)

        # Load detailed stats from fitzRoy CSVs (if available)
        detailed = load_player_detailed_stats(name)

        # Compute historical draft scores across career
        draft_history = compute_historical_draft_scores(player, detailed, players)

        # Load rating/potential from DB for this player
        from models.database import AflPlayer as _AP
        afl_row = _AP.query.filter_by(name=name).first()
        player_ratings = {}
        player_injury = {}
        if afl_row:
            player_ratings = {
                "rating": afl_row.rating,
                "potential": afl_row.potential,
            }
            if afl_row.injury_severity:
                from scrapers.afl_injuries import friendly_return_text
                from scrapers.squiggle import get_current_round
                current_round = get_current_round(config.CURRENT_YEAR)
                player_injury = {
                    "type": afl_row.injury_type,
                    "return": afl_row.injury_return,
                    "severity": afl_row.injury_severity,
                    "display": friendly_return_text(afl_row.injury_return, current_round),
                }

        # Acquisition info: find which league/team drafted this player
        acquisition_info = []
        if afl_row:
            from models.database import FantasyRoster, FantasyTeam, DraftPick, DraftSession, User, League
            acq_rows = (
                db.session.query(FantasyRoster, FantasyTeam)
                .join(FantasyTeam, FantasyRoster.team_id == FantasyTeam.id)
                .filter(FantasyRoster.player_id == afl_row.id, FantasyRoster.is_active == True)
                .all()
            )
            for fr, ft in acq_rows:
                owner = User.query.get(ft.owner_id) if ft.owner_id else None
                coach_name = (owner.display_name or owner.username) if owner else ft.name
                info = {
                    "coach": coach_name,
                    "team": ft.name,
                    "method": fr.acquired_via or "draft",
                    "league_name": "",
                }
                lg = db.session.get(League, ft.league_id)
                if lg:
                    info["league_name"] = lg.name
                # Find draft pick details
                if fr.acquired_via in ("draft", "supplemental", None):
                    dp = (
                        DraftPick.query
                        .join(DraftSession, DraftPick.draft_session_id == DraftSession.id)
                        .filter(
                            DraftSession.league_id == ft.league_id,
                            DraftPick.player_id == afl_row.id,
                            DraftPick.team_id == ft.id,
                        )
                        .first()
                    )
                    if dp:
                        ds = DraftSession.query.get(dp.draft_session_id)
                        info["pick_number"] = dp.pick_number
                        info["draft_year"] = ds.started_at.year if ds and ds.started_at else (ds.scheduled_start.year if ds and ds.scheduled_start else "")
                        info["draft_type"] = ds.draft_round_type if ds else "initial"
                acquisition_info.append(info)

        # State league data
        state_league_data = None
        if afl_row:
            from models.database import StateLeagueStat
            from sqlalchemy import func as sqla_func
            sl_rows = StateLeagueStat.query.filter(
                db.or_(
                    StateLeagueStat.player_id == afl_row.id,
                    StateLeagueStat.player_name == name,
                )
            ).order_by(StateLeagueStat.season.desc()).all()
            if sl_rows:
                def _build_section(rows):
                    career = []
                    for r in rows:
                        career.append({
                            "season": r.season, "competition": r.competition.upper(),
                            "team": r.team, "matches": r.matches, "age": r.age,
                            "disposals": r.disposals, "marks": r.marks, "goals": r.goals,
                            "tackles": r.tackles, "hitouts": r.hitouts,
                            "contested_possessions": r.contested_possessions,
                            "clearances": r.clearances, "inside_fifties": r.inside_fifties,
                            "fantasy_avg": r.dreamteam_avg,
                        })
                    latest = rows[0]
                    rankings = {}
                    if latest.dreamteam_avg and latest.competition and latest.season:
                        all_in_comp = StateLeagueStat.query.filter_by(
                            competition=latest.competition, season=latest.season
                        ).filter(StateLeagueStat.matches >= 3).all()
                        all_in_team = [s for s in all_in_comp if s.team == latest.team]
                        for label, pool, attr in [
                            ("comp", all_in_comp, "dreamteam_avg"),
                            ("team", all_in_team, "dreamteam_avg"),
                            ("comp_disposals", all_in_comp, "disposals"),
                            ("comp_tackles", all_in_comp, "tackles"),
                        ]:
                            vals = sorted([getattr(s, attr) or 0 for s in pool], reverse=True)
                            my_val = getattr(latest, attr) or 0
                            rank = next((i + 1 for i, v in enumerate(vals) if v <= my_val), len(vals))
                            rankings[label] = {"rank": rank, "of": len(vals), "value": round(my_val, 1)}
                    return {"career": career, "rankings": rankings,
                            "latest_comp": latest.competition.upper(),
                            "latest_season": latest.season}

                # Split into state league (VFL/SANFL/WAFL) and U18 (NAB/Coates)
                sl_senior = [r for r in sl_rows if r.competition != "nab"]
                sl_u18 = [r for r in sl_rows if r.competition == "nab"]
                if sl_senior:
                    state_league_data = _build_section(sl_senior)
                if sl_u18:
                    state_league_data = state_league_data or {}
                    state_league_data["u18"] = _build_section(sl_u18)

            # AFL team history — build from CSV data + state league listings
            afl_team_history = []
            from config import TEAM_LOGOS
            import pandas as pd
            import os as _os

            year_team = {}  # {year: (team_name, games_played)}
            data_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "data")

            # 1. Scan player_stats CSVs for actual team per year
            for yr in range(2013, config.CURRENT_YEAR + 1):
                csv_path = _os.path.join(data_dir, f"player_stats_{yr}.csv")
                if not _os.path.exists(csv_path):
                    continue
                try:
                    df = pd.read_csv(csv_path, low_memory=False)
                    if "Player" not in df.columns or "Team" not in df.columns:
                        continue
                    player_rows = df[df["Player"] == name]
                    if len(player_rows):
                        team = player_rows["Team"].mode().iloc[0] if len(player_rows) > 0 else None
                        if team:
                            year_team[yr] = (team, len(player_rows))
                except Exception:
                    pass

            # 2. Fill gaps from StateLeagueStat (AFL-listed but 0 AFL games)
            from models.database import StateLeagueStat as _SLS
            sl_listed = _SLS.query.filter(
                db.or_(_SLS.player_id == afl_row.id, _SLS.player_name == name),
                _SLS.is_afl_listed == True,
            ).all()
            for sl in sl_listed:
                if sl.season not in year_team:
                    # Listed at an AFL club but no AFL games — find which club
                    # Try to infer from adjacent years or current team
                    prev_team = year_team.get(sl.season - 1, (None, 0))[0]
                    next_team = year_team.get(sl.season + 1, (None, 0))[0]
                    team = prev_team or next_team or afl_row.afl_team
                    year_team[sl.season] = (team, 0)

            # 3. Current year — always show even if no games yet
            if config.CURRENT_YEAR not in year_team and afl_row.afl_team:
                year_team[config.CURRENT_YEAR] = (afl_row.afl_team, 0)

            # Build history list
            for yr in sorted(year_team.keys()):
                team, games = year_team[yr]
                afl_team_history.append({
                    "year": yr,
                    "team": team,
                    "logo": TEAM_LOGOS.get(team),
                    "games": games,
                })

        return render_template("player.html",
                               player=player,
                               breakdown=breakdown,
                               sc_history=sc_history,
                               detailed=detailed,
                               draft_history=draft_history,
                               weights=config.DRAFT_WEIGHTS,
                               player_ratings=player_ratings,
                               player_injury=player_injury,
                               acquisition_info=acquisition_info,
                               state_league_data=state_league_data,
                               afl_team_history=afl_team_history if afl_row else [])

    @app.route("/refresh", methods=["POST"])
    def trigger_refresh_all():
        """One-button refresh: rosters, SC scores, and fitzRoy detailed stats."""
        import pandas as pd

        steps = []

        # 1. Scrape rosters + current SC scores → master player list
        try:
            players = build_master_player_list()
            steps.append(f"Rosters: {len(players)} players")
        except Exception as e:
            steps.append(f"Rosters FAILED: {e}")

        # 2. Fetch detailed stats via fitzRoy (footywire source)
        script_path = os.path.join(config.BASE_DIR, "scripts", "fetch_stats.R")
        try:
            result = subprocess.run(
                [config.RSCRIPT_PATH, script_path, "2013", str(config.CURRENT_YEAR)],
                capture_output=True, text=True, timeout=600,
                cwd=config.BASE_DIR,
            )
            csv_count = sum(
                1 for y in range(2013, config.CURRENT_YEAR + 1)
                if os.path.exists(os.path.join(config.DATA_DIR, f"player_stats_{y}.csv"))
            )
            if result.returncode in (0, 127):  # 127 is OK on Windows bash
                steps.append(f"fitzRoy: {csv_count} years on disk")
            else:
                steps.append(f"fitzRoy error: {result.stderr[-200:]}")
        except FileNotFoundError:
            steps.append("fitzRoy: Rscript not found (install R + fitzRoy)")
        except subprocess.TimeoutExpired:
            steps.append("fitzRoy: timed out")
        except Exception as e:
            steps.append(f"fitzRoy: {e}")

        # 3. Clean names and filter detailed stats to current players only
        try:
            import re
            roster_df = pd.read_csv(os.path.join(config.DATA_DIR, "players.csv"))
            current_names = set(roster_df["name"].str.strip().str.lower())
            from scrapers.stats_loader import _NAME_FIXES
            trimmed = 0
            for year in range(2013, config.CURRENT_YEAR + 1):
                path = os.path.join(config.DATA_DIR, f"player_stats_{year}.csv")
                if not os.path.exists(path):
                    continue
                df = pd.read_csv(path)
                name_col = "Player" if "Player" in df.columns else None
                if name_col is None:
                    continue
                before = len(df)
                df[name_col] = df[name_col].apply(
                    lambda n: re.sub(r'\s*[\u2190-\u21ff\u2b00-\u2bff]+\s*$', '', str(n)).strip()
                    if isinstance(n, str) else str(n)
                )
                df[name_col] = df[name_col].apply(
                    lambda n: _NAME_FIXES.get(n.lower(), n)
                )
                df = df[df[name_col].str.strip().str.lower().isin(current_names)]
                df.to_csv(path, index=False)
                trimmed += before - len(df)
            steps.append(f"Filtered to current players (removed {trimmed:,} retired rows)")
        except Exception as e:
            steps.append(f"Filter: {e}")

        # 4. Backfill sc_avg from fitzRoy data for players missing it
        try:
            filled = backfill_sc_from_fitzroy()
            if filled > 0:
                steps.append(f"Backfilled SC avg for {filled} players from fitzRoy")
        except Exception as e:
            steps.append(f"SC backfill: {e}")

        # 5. Sync players.csv into AflPlayer DB table (force=True to bypass staleness)
        try:
            from models.database import AflPlayer
            _sync_players_to_db(app, force=True)
            total = AflPlayer.query.count()
            steps.append(f"DB sync: {total} players in AflPlayer table")
        except Exception as e:
            steps.append(f"DB sync FAILED: {e}")

        # 6. Re-sync ratings/potential from XLSX
        try:
            _sync_ratings_to_db(app)
            steps.append("Ratings synced from XLSX")
        except Exception as e:
            steps.append(f"Ratings sync: {e}")

        # 7. Pull AFL game schedule from Squiggle for current + next round
        try:
            from models.live_sync import sync_game_schedule
            from scrapers.squiggle import get_current_round

            year = config.CURRENT_YEAR
            current_round = get_current_round(year)
            if current_round is not None:
                count1 = sync_game_schedule(year, current_round)
                count2 = sync_game_schedule(year, current_round + 1)
                steps.append(f"Schedule: R{current_round}={count1} games, R{current_round+1}={count2} games")
            else:
                steps.append("Schedule: could not determine current round from Squiggle")
        except Exception as e:
            steps.append(f"Schedule sync: {e}")

        # 8. Ensure DB schema is up to date (idempotent)
        try:
            db.create_all()
            # Add indexes that create_all can't add to existing tables
            for idx_sql in [
                "CREATE INDEX IF NOT EXISTS ix_playerstat_year_round ON player_stat (year, round)",
                "CREATE INDEX IF NOT EXISTS ix_roster_team_active ON fantasy_roster (team_id, is_active)",
                "CREATE INDEX IF NOT EXISTS ix_fixture_league_round_year ON fixture (league_id, afl_round, year)",
                "CREATE INDEX IF NOT EXISTS ix_roundscore_round_year ON round_score (afl_round, year)",
            ]:
                try:
                    db.session.execute(db.text(idx_sql))
                except Exception:
                    pass
            db.session.commit()
            steps.append("Schema migration: OK")
        except Exception as e:
            steps.append(f"Schema migration: {e}")

        flash("Refresh complete — " + " | ".join(steps), "success")
        return redirect(url_for("admin.dashboard"))

    @app.route("/push/vapid-key")
    def vapid_key():
        return jsonify({"publicKey": config.VAPID_PUBLIC_KEY})

    @app.route("/push/subscribe", methods=["POST"])
    @login_required
    def push_subscribe():
        data = request.get_json(silent=True) or {}
        subscription = data.get("subscription")
        if subscription:
            import json as _json
            current_user.push_subscription = _json.dumps(subscription)
            db.session.commit()
        return jsonify({"ok": True})

    @app.route("/import/players", methods=["POST"])
    @login_required
    def trigger_import_players():
        if not getattr(current_user, "is_admin", False):
            flash("Admin access required.", "danger")
            return redirect(url_for("leagues.league_list"))
        f = request.files.get("file")
        if not f or f.filename == "":
            flash("No file uploaded.", "warning")
            return redirect(url_for("admin.dashboard"))
        filename = secure_filename(f.filename)
        if not filename.lower().endswith(".csv"):
            flash("Only CSV files are allowed.", "danger")
            return redirect(url_for("admin.dashboard"))
        path = os.path.join(config.DATA_DIR, "import_temp.csv")
        os.makedirs(config.DATA_DIR, exist_ok=True)
        f.save(path)
        try:
            players = import_players_csv(path, merge=True)
            flash(f"Imported/merged — master list now has {len(players)} players.", "success")
        except Exception as e:
            app.logger.exception("Player import failed")
            flash("Import failed. Check the CSV format and try again.", "danger")
        return redirect(url_for("admin.dashboard"))

    @app.route("/import/sc_scores", methods=["POST"])
    @login_required
    def trigger_import_sc():
        if not getattr(current_user, "is_admin", False):
            flash("Admin access required.", "danger")
            return redirect(url_for("leagues.league_list"))
        f = request.files.get("file")
        year = request.form.get("year", type=int) or config.CURRENT_YEAR
        if not f or f.filename == "":
            flash("No file uploaded.", "warning")
            return redirect(url_for("admin.dashboard"))
        filename = secure_filename(f.filename)
        if not filename.lower().endswith(".csv"):
            flash("Only CSV files are allowed.", "danger")
            return redirect(url_for("admin.dashboard"))
        path = os.path.join(config.DATA_DIR, "import_sc_temp.csv")
        os.makedirs(config.DATA_DIR, exist_ok=True)
        f.save(path)
        try:
            df = import_sc_scores_csv(path, year)
            flash(f"Imported {len(df)} SC score rows for {year}.", "success")
        except Exception as e:
            app.logger.exception("SC scores import failed")
            flash("Import failed. Check the CSV format and try again.", "danger")
        return redirect(url_for("admin.dashboard"))

    return app


# Create the app instance and SocketIO
app = create_app()

_allowed_origins = os.environ.get("ALLOWED_ORIGINS", "*")
if _allowed_origins != "*":
    _allowed_origins = [o.strip() for o in _allowed_origins.split(",")]
socketio = SocketIO(app, async_mode="threading", cors_allowed_origins=_allowed_origins,
                    ping_interval=25, ping_timeout=60)

# Exempt SocketIO from CSRF (it uses its own origin check)
csrf.exempt("flask_socketio")

# Register SocketIO events
from sockets.draft_events import register_draft_events
from sockets.matchup_events import register_matchup_events
from sockets.notification_events import register_notification_events
register_draft_events(socketio)
register_matchup_events(socketio)
register_notification_events(socketio)

# Give notification_manager access to socketio for real-time pushes
from models.notification_manager import init_notification_socketio
init_notification_socketio(socketio)

from models.activity_feed import init_activity_socketio
init_activity_socketio(socketio)

# Start live scoring scheduler (skip in testing)
if not app.testing:
    from models.scheduler import init_scheduler
    init_scheduler(app, socketio)

if __name__ == "__main__":
    os.makedirs(config.DATA_DIR, exist_ok=True)
    os.makedirs(os.path.join(config.DATA_DIR, "league"), exist_ok=True)
    debug = os.environ.get("FLASK_DEBUG", "1") == "1"
    port = int(os.environ.get("PORT", 5000))
    socketio.run(app, debug=debug, port=port)
